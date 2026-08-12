#!/usr/bin/env python3
"""Build row-level SIT contract-task ledger from the filtered PHU/SIT audit XLSX.

The output is used only for read-only drill-down in IT costs. It keeps source
accounting dimensions and validates that totals reconcile with contractTasks.json.
"""
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TARGET_TASKS = ("10", "22", "25")
MONTH_NAMES = [
    "január", "február", "marec", "apríl", "máj", "jún",
    "júl", "august", "september", "október", "november", "december",
]
REQUIRED_COLUMNS = {
    "Riadok", "Úloha", "Zákazka", "KPD", "PPD", "FZD", "PGD", "PRACM",
    "Suma", "Pôvodná ZAK", "Stĺpec", "Kategória", "Mesiac/obdobie", "Doklad",
    "Popis/poznámka", "Dátová poznámka",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_date(value: Any, year: int) -> tuple[str, int] | None:
    if isinstance(value, datetime):
        parsed = value
    else:
        raw = text(value)
        if not raw:
            return None
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return None
    if parsed.year != year:
        return None
    return parsed.date().isoformat(), parsed.month


def period_label(months_loaded: int, year: int) -> str:
    if months_loaded <= 0:
        return f"bez dát {year}"
    if months_loaded == 1:
        return f"{MONTH_NAMES[0]} {year}"
    return f"{MONTH_NAMES[0]} až {MONTH_NAMES[months_loaded - 1]} {year}"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, help="Filtrovaný audit XLSX")
    parser.add_argument("--tasks", type=Path, default=Path("src/data/contractTasks.json"))
    parser.add_argument("--output", type=Path, default=Path("src/data/contractTaskLedger.json"))
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()

    task_data = json.loads(args.tasks.read_text(encoding="utf-8"))
    months_loaded = int(task_data.get("meta", {}).get("monthsLoaded", 0))
    expected = {str(item["code"]): round(float(item["spent"]), 2) for item in task_data.get("tasks", [])}

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb["Filtrovaný audit"] if "Filtrovaný audit" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    missing = REQUIRED_COLUMNS.difference(headers)
    if missing:
        raise SystemExit(f"Missing XLSX columns: {', '.join(sorted(missing))}")
    index = {name: headers.index(name) for name in headers}

    ledger: list[dict[str, Any]] = []
    totals: dict[str, float] = defaultdict(float)
    month_totals: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    row_counts: dict[str, int] = defaultdict(int)

    for serial, row in enumerate(rows, start=2):
        task = text(row[index["Úloha"]])
        if task not in TARGET_TASKS:
            continue
        parsed = parse_date(row[index["Mesiac/obdobie"]], args.year)
        if parsed is None:
            continue
        date_iso, month = parsed
        if months_loaded and month > months_loaded:
            continue
        amount = round(float(row[index["Suma"]] or 0), 2)
        source_row = text(row[index["Riadok"]]) or str(serial)
        item = {
            "id": f"AUDIT-{source_row}",
            "sourceRow": source_row,
            "task": task,
            "zak": text(row[index["Zákazka"]]),
            "kpd": text(row[index["KPD"]]),
            "ppd": text(row[index["PPD"]]),
            "fzd": text(row[index["FZD"]]),
            "pgd": text(row[index["PGD"]]),
            "pracm": text(row[index["PRACM"]]),
            "amount": amount,
            "originalZak": text(row[index["Pôvodná ZAK"]]),
            "column": text(row[index["Stĺpec"]]),
            "category": text(row[index["Kategória"]]),
            "date": date_iso,
            "month": month,
            "document": text(row[index["Doklad"]]),
            "note": text(row[index["Popis/poznámka"]]),
            "dataNote": text(row[index["Dátová poznámka"]]),
        }
        ledger.append(item)
        totals[task] += amount
        month_totals[task][month] += amount
        row_counts[task] += 1

    ledger.sort(key=lambda item: (item["date"], item["task"], item["sourceRow"]))
    rounded_totals = {task: round(totals[task], 2) for task in TARGET_TASKS}
    for task in TARGET_TASKS:
        if task not in expected:
            raise SystemExit(f"Task {task} missing in {args.tasks}")
        if abs(rounded_totals[task] - expected[task]) > 0.01:
            raise SystemExit(
                f"Reconciliation failed for task {task}: ledger={rounded_totals[task]:.2f}, "
                f"contractTasks={expected[task]:.2f}"
            )

    result = {
        "meta": {
            "title": "SIT 2026 – riadkový drill-down kontraktových úloh",
            "source": "Filtrovaný audit PHU/SIT",
            "sourceFile": args.xlsx.name,
            "period": period_label(months_loaded, args.year),
            "year": args.year,
            "monthsLoaded": months_loaded,
            "rowCount": len(ledger),
            "rowCountByTask": {task: row_counts[task] for task in TARGET_TASKS},
            "totalsByTask": rounded_totals,
            "method": (
                "Riadky sú prevzaté bez dopočtu z Filtrovaného auditu. Výber obsahuje iba úlohy 10, 22 a 25. "
                "Súčet riadkov je pri generovaní povinne reconciliovaný s contractTasks.json."
            ),
        },
        "payments": ledger,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Written {args.output} ({len(ledger)} rows)")
    for task in TARGET_TASKS:
        monthly = ", ".join(f"{month}:{month_totals[task][month]:.2f}" for month in sorted(month_totals[task]))
        print(f"Task {task}: rows={row_counts[task]}, total={rounded_totals[task]:.2f}, monthly=[{monthly}]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
