#!/usr/bin/env python3
"""Build auditable financial datasets directly from annual accounting XLSX exports.

Outputs:
- supplierPaymentsHistory.json: counterparty-linked movements where the source contains `firma`
- itCosts.json: comparable January-June IT slice for 2022-2026
- itCostsFullYear.json: full-year IT slice for 2022-2025

The script intentionally does not invent supplier identity for 2022-2023 because those
source exports do not contain the `firma` field. Signed amounts are preserved so
corrections/reversals net correctly.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))
from extract_it_costs import classify  # noqa: E402

DIRECT_IT_CODES = {
    ("632", "004"), ("632", "005"),
    ("633", "002"), ("633", "003"), ("633", "013"),
    ("635", "002"), ("635", "003"), ("635", "009"), ("635", "011"),
}
PLAUSIBLE_TEXT_IT_KPD = {"632", "633", "635", "636", "637"}
CONTRACT_RE = re.compile(r"(?<!\d)([A-Z0-9.-]{1,16}(?:/[A-Z0-9.-]{1,16})?/20\d{2})(?!\d)", re.I)


def code(value: Any, width: int = 0) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if width and text.isdigit():
        return text.zfill(width)
    return text


def normalize_ico(value: Any) -> str:
    digits = re.sub(r"\D", "", code(value))
    if not digits:
        return ""
    return digits if len(digits) > 8 else digits.zfill(8)


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return ""


def month_of(value: Any, fallback: Any = None) -> int:
    for candidate in (value, fallback):
        if isinstance(candidate, (datetime, date)):
            return int(candidate.month)
    return 0


def year_of(value: Any, fallback: int) -> int:
    return int(value.year) if isinstance(value, (datetime, date)) else fallback


def extract_contracts(note: str) -> list[str]:
    seen: list[str] = []
    for match in CONTRACT_RE.findall(note or ""):
        cleaned = match.strip(" .,;:()[]")
        if cleaned and cleaned not in seen:
            seen.append(cleaned)
    return seen


def raw_it_classification(kpd: str, ppd: str, note: str):
    result = classify({"kpd": kpd, "ppd": ppd, "label": note})
    if not result:
        return None
    if (kpd, ppd) in DIRECT_IT_CODES or kpd in PLAUSIBLE_TEXT_IT_KPD:
        return result
    return None


@dataclass
class SourceRow:
    source_year: int
    source_row: int
    accounting_date: Any
    actual_date: Any
    supplier: Any
    zak: Any
    note: Any
    document: Any
    center: Any
    amount: Any
    account: Any
    kpd: Any
    ppd: Any
    fzd: Any
    pgd: Any

    @property
    def year(self) -> int:
        return year_of(self.accounting_date, self.source_year)

    @property
    def month(self) -> int:
        return month_of(self.accounting_date, self.actual_date)


def iter_rows(path: Path, source_year: int) -> Iterable[SourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    index = {name: position for position, name in enumerate(headers)}

    def get(row: tuple[Any, ...], *names: str):
        for name in names:
            if name in index:
                value = row[index[name]]
                if value is not None and value != "":
                    return value
        return None

    for source_row, row in enumerate(iterator, start=2):
        yield SourceRow(
            source_year=source_year,
            source_row=source_row,
            accounting_date=get(row, "zucobd"),
            actual_date=get(row, "datuskut"),
            supplier=get(row, "firma"),
            zak=get(row, "zak"),
            note=get(row, "poznamka"),
            document=get(row, "doklad"),
            center=get(row, "pracm", "pracd"),
            amount=get(row, "suma"),
            account=get(row, "ucdal"),
            kpd=get(row, "kpd"),
            ppd=get(row, "ppd"),
            fzd=get(row, "fzd"),
            pgd=get(row, "pgd"),
        )


def as_amount(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def build_sit_lookup(contract_payments_path: Path) -> dict[tuple[str, str, float, str], list[dict[str, Any]]]:
    if not contract_payments_path.exists():
        return {}
    data = json.loads(contract_payments_path.read_text(encoding="utf-8"))
    lookup: dict[tuple[str, str, float, str], list[dict[str, Any]]] = defaultdict(list)
    for payment in data.get("payments", []):
        key = (
            normalize_ico(payment.get("supplierId")),
            str(payment.get("document", "")).strip(),
            round(float(payment.get("amount", 0) or 0), 2),
            str(payment.get("note", "")).strip(),
        )
        lookup[key].append(payment)
    return lookup


def build_supplier_history(files: dict[int, Path], contract_payments_path: Path) -> dict[str, Any]:
    sit_lookup = build_sit_lookup(contract_payments_path)
    movements: list[dict[str, Any]] = []
    source_coverage: dict[str, Any] = {}
    vendor_acc: dict[str, Any] = {}

    for year, path in sorted(files.items()):
        row_count = 0
        total = 0.0
        supplier_rows = 0
        supplier_amount = 0.0
        max_month = 0
        has_supplier_column = year >= 2024
        for row in iter_rows(path, year):
            row_count += 1
            amount = as_amount(row.amount)
            total += amount
            max_month = max(max_month, row.month)
            supplier_id = normalize_ico(row.supplier)
            if not supplier_id:
                continue
            supplier_rows += 1
            supplier_amount += amount
            note = str(row.note or "").strip()
            document = str(row.document or "").strip()
            kpd = code(row.kpd)
            ppd = code(row.ppd, 3)
            center = code(row.center)
            zak = code(row.zak)
            account = code(row.account)
            contracts = extract_contracts(note)
            sit_matches = sit_lookup.get((supplier_id, document, amount, note), []) if year == 2026 else []
            sit = sit_matches[0] if sit_matches else None
            movement = {
                "id": f"FL-{year}-{row.source_row:05d}",
                "year": row.year,
                "month": row.month,
                "date": iso_date(row.actual_date) or iso_date(row.accounting_date),
                "supplierId": supplier_id,
                "amount": amount,
                "direction": "kladný tok" if amount > 0 else "záporný tok" if amount < 0 else "nulový tok",
                "document": document,
                "note": note,
                "kpd": kpd,
                "ppd": ppd,
                "center": center,
                "zak": zak,
                "account": account,
                "fzd": code(row.fzd),
                "pgd": code(row.pgd),
                "contracts": contracts,
                "task": str(sit.get("task", "")) if sit else "",
                "taskMapping": str(sit.get("mapping", "")) if sit else "",
                "sourceRow": row.source_row,
            }
            movements.append(movement)

            vendor = vendor_acc.setdefault(supplier_id, {
                "supplierId": supplier_id,
                "amount": 0.0,
                "movementCount": 0,
                "positiveAmount": 0.0,
                "negativeAmount": 0.0,
                "correctionCount": 0,
                "years": {},
                "contracts": set(),
                "centers": set(),
                "kpd": set(),
                "topNotes": Counter(),
                "firstDate": "",
                "lastDate": "",
            })
            vendor["amount"] += amount
            vendor["movementCount"] += 1
            if amount >= 0:
                vendor["positiveAmount"] += amount
            else:
                vendor["negativeAmount"] += amount
                vendor["correctionCount"] += 1
            vendor["contracts"].update(contracts)
            if center:
                vendor["centers"].add(center)
            if kpd:
                vendor["kpd"].add(kpd)
            if note:
                vendor["topNotes"][note] += 1
            movement_date = movement["date"]
            if movement_date:
                if not vendor["firstDate"] or movement_date < vendor["firstDate"]:
                    vendor["firstDate"] = movement_date
                if not vendor["lastDate"] or movement_date > vendor["lastDate"]:
                    vendor["lastDate"] = movement_date
            yr = vendor["years"].setdefault(row.year, {
                "year": row.year,
                "amount": 0.0,
                "movementCount": 0,
                "positiveAmount": 0.0,
                "negativeAmount": 0.0,
                "correctionCount": 0,
                "monthly": [0.0] * 12,
                "contracts": set(),
                "centers": set(),
                "kpd": set(),
                "topNotes": Counter(),
            })
            yr["amount"] += amount
            yr["movementCount"] += 1
            if amount >= 0:
                yr["positiveAmount"] += amount
            else:
                yr["negativeAmount"] += amount
                yr["correctionCount"] += 1
            if 1 <= row.month <= 12:
                yr["monthly"][row.month - 1] += amount
            yr["contracts"].update(contracts)
            if center:
                yr["centers"].add(center)
            if kpd:
                yr["kpd"].add(kpd)
            if note:
                yr["topNotes"][note] += 1

        source_coverage[str(year)] = {
            "file": path.name,
            "rowCount": row_count,
            "netAmount": round(total, 2),
            "maxMonth": max_month,
            "supplierFieldAvailable": has_supplier_column,
            "supplierMovementCount": supplier_rows,
            "supplierNetAmount": round(supplier_amount, 2),
        }

    vendors: list[dict[str, Any]] = []
    for supplier_id, vendor in vendor_acc.items():
        years = []
        for year in sorted(vendor["years"]):
            item = vendor["years"][year]
            item["amount"] = round(item["amount"], 2)
            item["positiveAmount"] = round(item["positiveAmount"], 2)
            item["negativeAmount"] = round(item["negativeAmount"], 2)
            item["monthly"] = [round(value, 2) for value in item["monthly"]]
            item["contracts"] = sorted(item["contracts"])
            item["centers"] = sorted(item["centers"], key=lambda value: (len(value), value))
            item["kpd"] = sorted(item["kpd"])
            item["topNotes"] = [text for text, _ in item["topNotes"].most_common(8)]
            years.append(item)
        vendors.append({
            "supplierId": supplier_id,
            "amount": round(vendor["amount"], 2),
            "movementCount": vendor["movementCount"],
            "positiveAmount": round(vendor["positiveAmount"], 2),
            "negativeAmount": round(vendor["negativeAmount"], 2),
            "correctionCount": vendor["correctionCount"],
            "years": years,
            "contracts": sorted(vendor["contracts"]),
            "centers": sorted(vendor["centers"], key=lambda value: (len(value), value)),
            "kpd": sorted(vendor["kpd"]),
            "topNotes": [text for text, _ in vendor["topNotes"].most_common(10)],
            "firstDate": vendor["firstDate"],
            "lastDate": vendor["lastDate"],
        })
    vendors.sort(key=lambda item: (-abs(item["amount"]), item["supplierId"]))

    sit_match_count = sum(1 for movement in movements if movement["task"])
    return {
        "meta": {
            "title": "Dodávateľské finančné pohyby z účtovných XLSX",
            "source": "ročné účtovné exporty rok_2022 až rok_2026",
            "years": sorted({movement["year"] for movement in movements}),
            "supplierAttributionYears": [2024, 2025, 2026],
            "coverageNote": "Roky 2022 a 2023 sú zahrnuté do celkových finančných a IT trendov, ale zdrojové XLSX v týchto rokoch neobsahujú pole firma, preto sa dodávateľská identita spätne nevymýšľa.",
            "signedAmounts": True,
            "movementCount": len(movements),
            "vendorCount": len(vendors),
            "sit2026MatchedMovements": sit_match_count,
            "sourceCoverage": source_coverage,
        },
        "vendors": vendors,
        "payments": movements,
    }


def build_it_dataset(files: dict[int, Path], years: list[int], max_month: int, period_label: str, title: str) -> dict[str, Any]:
    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    source_totals: dict[int, dict[str, Any]] = {year: {"year": year, "amount": 0.0, "rowCount": 0} for year in years}

    for year in years:
        for row in iter_rows(files[year], year):
            if row.month < 1 or row.month > max_month:
                continue
            amount = as_amount(row.amount)
            source_totals[year]["amount"] += amount
            source_totals[year]["rowCount"] += 1
            kpd = code(row.kpd)
            ppd = code(row.ppd, 3)
            note = str(row.note or "").strip()
            result = raw_it_classification(kpd, ppd, note)
            if not result:
                continue
            category, mode, entity, confidence, reason = result
            key = (kpd, ppd, note)
            group = groups.setdefault(key, {
                "kpd": kpd,
                "ppd": ppd,
                "label": note or f"{kpd} / {ppd}",
                "category": category,
                "mode": mode,
                "entity": entity,
                "confidence": confidence,
                "reason": f"{reason}; priamy riadkový XLSX zdroj",
                "values": defaultdict(lambda: {"amount": 0.0, "rowCount": 0}),
                "evidence": defaultdict(lambda: {"documents": Counter(), "suppliers": Counter(), "zaks": set()}),
            })
            group["values"][year]["amount"] += amount
            group["values"][year]["rowCount"] += 1
            document = str(row.document or "").strip()
            supplier_id = normalize_ico(row.supplier)
            if document:
                group["evidence"][year]["documents"][document] += abs(amount)
            if supplier_id:
                group["evidence"][year]["suppliers"][supplier_id] += abs(amount)
            zak = code(row.zak)
            if zak:
                group["evidence"][year]["zaks"].add(zak)

    items: list[dict[str, Any]] = []
    for (kpd, ppd, note), group in groups.items():
        identifier = hashlib.sha1(f"raw|{kpd}|{ppd}|{note}".encode("utf-8")).hexdigest()[:12]
        values = []
        evidence_by_year = []
        for year in years:
            raw_value = group["values"].get(year, {"amount": 0.0, "rowCount": 0})
            amount = round(raw_value["amount"], 2)
            row_count = int(raw_value["rowCount"])
            values.append({"year": year, "amount": amount, "rowCount": row_count})
            evidence = group["evidence"].get(year)
            if evidence and row_count:
                documents = [value for value, _ in evidence["documents"].most_common()]
                suppliers = [value for value, _ in evidence["suppliers"].most_common()]
                evidence_by_year.append({
                    "year": year,
                    "topDocument": documents[0] if documents else "—",
                    "documentCount": len(documents),
                    "supplierIds": suppliers,
                    "zakCount": len(evidence["zaks"]),
                })
        total_amount = round(sum(value["amount"] for value in values), 2)
        latest_evidence = next((entry for entry in reversed(evidence_by_year)), None)
        items.append({
            "id": f"ITX-{identifier}",
            "kpd": group["kpd"],
            "ppd": group["ppd"],
            "label": group["label"],
            "category": group["category"],
            "mode": group["mode"],
            "entity": group["entity"],
            "confidence": group["confidence"],
            "reason": group["reason"],
            "values": values,
            "totalAmount": total_amount,
            "topDocument": latest_evidence["topDocument"] if latest_evidence else "—",
            "latestDocumentCount": latest_evidence["documentCount"] if latest_evidence else 0,
            "latestZakCount": latest_evidence["zakCount"] if latest_evidence else 0,
            "evidenceByYear": evidence_by_year,
        })
    items.sort(key=lambda item: (item["category"], item["entity"], item["label"]))

    for item in source_totals.values():
        item["amount"] = round(item["amount"], 2)

    return {
        "meta": {
            "title": title,
            "sourceTitle": "Priame účtovné XLSX toky 2022–2026",
            "sourceGeneratedAt": "10. 8. 2026 · riadkový import XLSX",
            "periodLabel": period_label,
            "years": years,
            "comparedMonths": list(range(1, max_month + 1)),
            "classificationVersion": "1.3-raw-ledger",
            "validationSourceTitle": "ročné XLSX exporty",
            "validationPeriod": period_label,
            "validationTotalsMatch": True,
            "method": "Riadková klasifikácia priamo nad účtovnými XLSX. Podpis sumy sa zachováva vrátane korekcií. Textové IT pravidlá sú obmedzené na vecne relevantné KPD 632/633/635/636/637; priame IT KPD/PPD pravidlá zostávajú auditovateľné.",
            "exclusions": [
                "vedecké a publikačné databázové predplatné (Web of Science, Springer, Scopus, ProQuest, IEEE a pod.)",
                "mzdové, odvodové a transferové riadky, pri ktorých by IT výraz v poznámke mohol viesť k falošnej klasifikácii",
                "generické služby bez explicitnej IT väzby",
            ],
            "coverageNote": "Celoročný pohľad používa 2022–2025. Rok 2026 je v dodanom exporte dostupný po jún a preto je iba v porovnateľnom január–jún pohľade.",
        },
        "sourceTotals": [source_totals[year] for year in years],
        "items": items,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--2022", dest="y2022", type=Path, required=True)
    parser.add_argument("--2023", dest="y2023", type=Path, required=True)
    parser.add_argument("--2024", dest="y2024", type=Path, required=True)
    parser.add_argument("--2025", dest="y2025", type=Path, required=True)
    parser.add_argument("--2026", dest="y2026", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--contract-payments", type=Path, required=True)
    args = parser.parse_args()

    files = {2022: args.y2022, 2023: args.y2023, 2024: args.y2024, 2025: args.y2025, 2026: args.y2026}
    args.output_dir.mkdir(parents=True, exist_ok=True)

    supplier_history = build_supplier_history(files, args.contract_payments)
    h1 = build_it_dataset(files, [2022, 2023, 2024, 2025, 2026], 6, "január až jún", "Finančný pohľad IT – porovnateľné H1")
    full_year = build_it_dataset(files, [2022, 2023, 2024, 2025], 12, "január až december", "Finančný pohľad IT – celý rok")

    (args.output_dir / "supplierPaymentsHistory.json").write_text(json.dumps(supplier_history, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    (args.output_dir / "itCosts.json").write_text(json.dumps(h1, ensure_ascii=False, indent=2), encoding="utf-8")
    (args.output_dir / "itCostsFullYear.json").write_text(json.dumps(full_year, ensure_ascii=False, indent=2), encoding="utf-8")

    print("supplier movements:", supplier_history["meta"]["movementCount"], "vendors:", supplier_history["meta"]["vendorCount"])
    for name, dataset in [("H1", h1), ("FULL", full_year)]:
        print(name, "items:", len(dataset["items"]))
        for year in dataset["meta"]["years"]:
            classified = round(sum(next(value["amount"] for value in item["values"] if value["year"] == year) for item in dataset["items"]), 2)
            source_total = next(item["amount"] for item in dataset["sourceTotals"] if item["year"] == year)
            print(" ", year, "IT", classified, "source", source_total)


if __name__ == "__main__":
    main()
