from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
import sys
if len(sys.argv) < 2:
    raise SystemExit('Použitie: python scripts/build_seed.py /cesta/k/Navrh_RACI.xlsx')
XLSX = Path(sys.argv[1]).expanduser().resolve()
if not XLSX.exists():
    raise SystemExit(f'Vstupný súbor neexistuje: {XLSX}')


def clean(value: Any) -> Any:
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass
    if isinstance(value, str) and value.startswith('='):
        return ''
    return value


def slugify(text: str) -> str:
    normalized = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    normalized = re.sub(r'[^a-zA-Z0-9]+', '-', normalized).strip('-').lower()
    return normalized


def rows(ws_name: str):
    ws = wb[ws_name]
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[4]]
    out = []
    for values in ws.iter_rows(min_row=5, values_only=True):
        if not any(v is not None for v in values):
            continue
        out.append({headers[i]: clean(v) for i, v in enumerate(values) if i < len(headers) and headers[i]})
    return out


wb = load_workbook(XLSX, data_only=False)

role_rows = rows('Role a kompetencie')
employees = []
for r in role_rows:
    name = r['Zamestnanec']
    if not name:
        continue
    employees.append({
        'id': slugify(name),
        'name': name,
        'position': r.get('Formálna pozícia', ''),
        'roleType': r.get('Typ roly', ''),
        'responsibilities': r.get('Hlavné zodpovednosti', ''),
        'systems': r.get('Systémy / projekty', ''),
        'decides': r.get('Samostatne rozhoduje o', ''),
        'needsApproval': r.get('Vyžaduje schválenie pri', ''),
        'outputs': r.get('Hlavné výstupy', ''),
        'manager': r.get('Priamy nadriadený', ''),
        'deputy': r.get('Zástupca', ''),
        'capacity': r.get('Kapacita %', ''),
        'documentation': r.get('Dokumentácia', ''),
        'status': r.get('Stav potvrdenia', '') or 'Na potvrdenie',
        'note': r.get('Poznámka', ''),
    })

raci_rows = rows('RACI')
participant_headers = [
    'Pavol Horváth', 'Peter Modrák', 'Ladislav Turányi', 'Miroslav Kozel', 'Martin Vozák',
    'Csongor Mészáros', 'Martin Korének', 'Roman Vápeník', 'Dávid Cymbalák',
    'Lukáš Visokai', 'Michelle Kožuchová Bajema', 'Vecný garant / MŠVVaM',
    'Iné útvary CVTI SR', 'Dodávateľ / partner'
]
raci = []
for r in raci_rows:
    assignments = {p: r.get(p, '') or '' for p in participant_headers}
    raci.append({
        'id': r['ID'],
        'area': r['Oblasť'],
        'process': r['Proces / služba / rozhodnutie'],
        'output': r['Hlavný výstup'],
        'criticality': r['Kritickosť'],
        'assignments': assignments,
        'note': r.get('Poznámka / čo potvrdiť', ''),
    })

service_rows = rows('Kritické služby')
services = []
for r in service_rows:
    services.append({
        'id': r['ID'],
        'name': r['Služba / systém'],
        'category': r.get('Kategória', ''),
        'criticality': r.get('Kritickosť', ''),
        'businessOwner': r.get('Biznis vlastník', ''),
        'technicalOwner': r.get('Technický vlastník', ''),
        'primary': r.get('Primárny vykonávateľ', ''),
        'deputy': r.get('Zástupca', ''),
        'rto': r.get('Max. tolerovaný výpadok / RTO', ''),
        'runbook': r.get('Dokumentácia / runbook', ''),
        'repository': r.get('Repozitár / zdrojové kódy', ''),
        'monitoring': r.get('Monitoring', ''),
        'backup': r.get('Zálohovanie a obnova', ''),
        'securityOwner': r.get('Bezpečnostný vlastník', ''),
        'supplierSla': r.get('Dodávateľ / SLA', ''),
        'readiness': r.get('Stav pripravenosti', '') or 'Na potvrdenie',
        'note': r.get('Poznámka', ''),
    })

sub_rows = rows('Zastupiteľnosť')
substitutions = []
for r in sub_rows:
    substitutions.append({
        'id': r['ID'],
        'agenda': r['Kritická služba / agenda'],
        'owner': r.get('Primárny vlastník', ''),
        'currentState': r.get('Súčasný stav', ''),
        'proposedDeputy': r.get('Navrhovaný zástupca – pracovný návrh', ''),
        'confirmedDeputy': r.get('Potvrdený zástupca', ''),
        'scope': r.get('Rozsah zastúpenia', ''),
        'runbook': r.get('Minimálna dokumentácia / runbook', ''),
        'location': r.get('Miesto dokumentácie', ''),
        'handoverDue': r.get('Termín odovzdania', ''),
        'testDate': r.get('Dátum testu zastúpenia', ''),
        'testResult': r.get('Výsledok testu', ''),
        'status': r.get('Stav', '') or 'Na potvrdenie',
        'note': r.get('Poznámka', ''),
    })

cap_rows = rows('Kapacity')
capacity = []
for r in cap_rows:
    capacity.append({
        'employee': r['Zamestnanec'],
        'management': r.get('Riadenie / administratíva %', '') or 0,
        'operations': r.get('Bežná prevádzka %', '') or 0,
        'projects': r.get('Projekty / rozvoj %', '') or 0,
        'helpdesk': r.get('Helpdesk / incidenty %', '') or 0,
        'other': r.get('Iné činnosti %', '') or 0,
        'seasonalPeaks': r.get('Sezónne špičky', ''),
        'conflict': r.get('Najväčší kapacitný konflikt', ''),
        'sustainability': r.get('Nadčasy / udržateľnosť', ''),
        'status': r.get('Stav', '') or 'Na potvrdenie',
        'note': r.get('Poznámka zo zdroja', ''),
    })

risk_rows = rows('Riziká')
risks = []
for r in risk_rows:
    probability = int(r.get('Pravdepodobnosť 1–5', 0) or 0)
    impact_score = int(r.get('Dopad 1–5', 0) or 0)
    score = probability * impact_score
    priority = 'Kritická' if score >= 20 else 'Vysoká' if score >= 12 else 'Stredná' if score >= 6 else 'Nízka'
    risks.append({
        'id': r['ID'],
        'area': r['Oblasť'],
        'risk': r['Riziko'],
        'trigger': r.get('Spúšťač / príčina', ''),
        'impact': r.get('Dopad', ''),
        'probability': probability,
        'impactScore': impact_score,
        'priority': priority,
        'owner': r.get('Vlastník rizika', ''),
        'measure': r.get('Navrhované opatrenie', ''),
        'due': r.get('Termín', ''),
        'status': r.get('Stav', '') or 'Otvorené',
        'evidence': r.get('Dôkaz / odkaz', ''),
        'managementDecision': r.get('Rozhodnutie vedenia', ''),
        'note': r.get('Poznámka', ''),
    })

action_rows = rows('Akčný plán')
actions = []
for r in action_rows:
    actions.append({
        'id': r['ID'],
        'horizon': r.get('Horizont', ''),
        'title': r['Krok / opatrenie'],
        'expectedOutput': r.get('Očakávaný výstup', ''),
        'proposedOwner': r.get('Navrhovaný vlastník', ''),
        'confirmedOwner': r.get('Potvrdený vlastník', ''),
        'start': r.get('Začiatok', ''),
        'due': r.get('Termín', ''),
        'status': r.get('Stav', '') or 'Návrh',
        'dependency': r.get('Závislosť', ''),
        'kpi': r.get('KPI / kritérium splnenia', ''),
        'directorDecision': r.get('Rozhodnutie riaditeľa', ''),
        'note': r.get('Poznámka', ''),
    })

decision_rows = rows('Rozhodnutia')
decisions = []
for r in decision_rows:
    decisions.append({
        'id': r['ID'],
        'topic': r['Téma'],
        'question': r['Otázka na rozhodnutie'],
        'proposal': r.get('Pracovný návrh', ''),
        'decisionMaker': r.get('Rozhodovateľ', ''),
        'due': r.get('Termín', ''),
        'decision': r.get('Rozhodnutie', ''),
        'reason': r.get('Dôvod / podklad', ''),
        'impact': r.get('Dopad rozhodnutia', ''),
        'status': r.get('Stav', '') or 'Návrh',
        'note': r.get('Poznámka', ''),
    })

projects = [
    {'id':'P01','name':'KOMIS','type':'Program / IS','owner':'Peter Modrák','sponsor':'Pavol Horváth','status':'Prebieha','priority':'Vysoká','progress':35,'start':'','due':'','description':'Koordinácia zmien a projektových úloh v rámci KOMIS.'},
    {'id':'P02','name':'MOHOK','type':'Projekt / web','owner':'Peter Modrák','sponsor':'Pavol Horváth','status':'Prebieha','priority':'Vysoká','progress':45,'start':'','due':'','description':'Koordinácia a technické práce na projekte MOHOK.'},
    {'id':'P03','name':'ISOVAV','type':'Projekt','owner':'Peter Modrák','sponsor':'Pavol Horváth','status':'Plánované','priority':'Stredná','progress':20,'start':'','due':'','description':'Projektové riadenie a odborné výstupy ISOVAV.'},
    {'id':'P04','name':'Nový intranet','type':'Aplikačný rozvoj','owner':'Csongor Mészáros','sponsor':'Peter Modrák','status':'Prebieha','priority':'Kritická','progress':60,'start':'','due':'','description':'Vývoj, nasadenie a príprava prevádzkového modelu nového intranetu.'},
    {'id':'P05','name':'InnovAIte','type':'Medzinárodný projekt','owner':'Dávid Cymbalák','sponsor':'Pavol Horváth','status':'Prebieha','priority':'Vysoká','progress':50,'start':'','due':'','description':'Projektové, infraštruktúrne a partnerské výstupy InnovAIte.'},
    {'id':'P06','name':'EOSC služby','type':'Program / služby','owner':'Ladislav Turányi','sponsor':'Pavol Horváth','status':'Na rozhodnutie','priority':'Vysoká','progress':25,'start':'','due':'','description':'Nastavenie cieľového modelu EOSC aktivít, portálov a katalógu.'},
    {'id':'P07','name':'Obnova NTI','type':'Prevádzkový program','owner':'Ladislav Turányi','sponsor':'Pavol Horváth','status':'Na rozhodnutie','priority':'Kritická','progress':15,'start':'','due':'','description':'Cieľový stav NTI, licencie, zariadenia, miestnosti a serverové časti.'},
    {'id':'P08','name':'NISPEZ V','type':'Projekt','owner':'Michelle Kožuchová Bajema','sponsor':'Peter Modrák','status':'Prebieha','priority':'Stredná','progress':40,'start':'','due':'','description':'Grafické a publikačné projektové výstupy.'},
]

tasks = []
status_map = {'Návrh':'Návrh','Rozpracované':'Prebieha','Ukončené':'Hotovo'}
for i, a in enumerate(actions, start=1):
    tasks.append({
        'id': f'T{i:02d}',
        'title': a['title'],
        'projectId': '',
        'owner': a['confirmedOwner'] or a['proposedOwner'],
        'priority': 'Vysoká' if i <= 8 else 'Stredná',
        'status': status_map.get(a['status'], 'Návrh'),
        'due': a['due'],
        'description': a['expectedOutput'],
        'source': a['id'],
    })

seed = {
    'meta': {
        'version': '0.1.0',
        'organization': 'CVTI SR',
        'unit': 'Odbor 3.2',
        'sourceDate': '2026-07-30',
        'source': 'RACI a manažérsky popis fungovania odboru',
    },
    'employees': employees,
    'raci': raci,
    'services': services,
    'substitutions': substitutions,
    'capacity': capacity,
    'risks': risks,
    'actions': actions,
    'decisions': decisions,
    'projects': projects,
    'tasks': tasks,
}

out = ROOT / 'src' / 'data' / 'seed.json'
out.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding='utf-8')
print(out)
print({k: len(v) for k, v in seed.items() if isinstance(v, list)})
