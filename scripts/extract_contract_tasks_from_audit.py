#!/usr/bin/env python3
"""Regenerate src/data/contractTasks.json from the filtered PHU/SIT audit XLSX.

Only tasks 10, 22 and 25 are aggregated. Task names, descriptions, centers and
budgets are preserved from the existing contractTasks.json, while monthly spend,
spent total and remaining budget are recalculated from the source audit.
"""
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TARGET_TASKS = ("10", "22", "25")
MONTH_NAMES = [
    "január", "február", "marec", "apríl", "máj", "jún",
    "júl", "august", "september", "október", "november", "december",
]
REQUIRED_COLUMNS = {"Úloha", "Suma", "Mesiac/obdobie"}


def parse_month(value: Any, year: int) -> int | None:
    if isinstance(value, datetime):
        return value.month if value.year == year else None
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.month if parsed.year == year else None


def period_label(months_loaded: int, year: int) -> str:
    if months_loaded <= 0:
        return f"bez dát {year}"
    if months_loaded == 1:
        return f"{MONTH_NAMES[0]} {year}"
    return f"{MONTH_NAMES[0]} až {MONTH_NAMES[months_loaded - 1]} {year}"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, help="Filtrovaný audit XLSX")
    parser.add_argument("--base", type=Path, default=Path("src/data/contractTasks.json"))
    parser.add_argument("--output", type=Path, default=Path("src/data/contractTasks.json"))
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--months", type=int, default=0, help="Expected loaded months; 0 = infer from data")
    args = parser.parse_args()

    base = json.loads(args.base.read_text(encoding="utf-8"))
    base_by_code = {str(task["code"]): task for task in base.get("tasks", [])}
    missing_base = [code for code in TARGET_TASKS if code not in base_by_code]
    if missing_base:
        raise SystemExit(f"Missing base task definitions: {', '.join(missing_base)}")

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb["Filtrovaný audit"] if "Filtrovaný audit" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    missing = REQUIRED_COLUMNS.difference(headers)
    if missing:
        raise SystemExit(f"Missing XLSX columns: {', '.join(sorted(missing))}")
    index = {name: headers.index(name) for name in headers}

    monthly: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    row_counts: dict[str, int] = defaultdict(int)
    centers: dict[str, set[str]] = defaultdict(set)
    max_month = 0
    for row in rows:
        code = str(row[index["Úloha"]] or "").strip()
        if code not in TARGET_TASKS:
            continue
        month = parse_month(row[index["Mesiac/obdobie"]], args.year)
        if month is None:
            continue
        amount = float(row[index["Suma"]] or 0)
        monthly[code][month] += amount
        row_counts[code] += 1
        if "PRACM" in index:
            center = str(row[index["PRACM"]] or "").strip()
            if center:
                centers[code].add(center)
        max_month = max(max_month, month)

    months_loaded = args.months or max_month
    if not 1 <= months_loaded <= 12:
        raise SystemExit(f"Invalid monthsLoaded={months_loaded}")
    if max_month > months_loaded:
        raise SystemExit(f"Source contains month {max_month}, beyond requested --months {months_loaded}")

    tasks = []
    for code in TARGET_TASKS:
        original = base_by_code[code]
        values = [round(monthly[code].get(month, 0.0), 2) for month in range(1, months_loaded + 1)]
        spent = round(sum(values), 2)
        budget = round(float(original["budget"]), 2)
        task = dict(original)
        if centers[code]:
            task["centers"] = sorted(centers[code], key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value))
        task["spent"] = spent
        task["remaining"] = round(budget - spent, 2)
        task["monthly"] = values
        tasks.append(task)

    result = {
        "meta": {
            "title": "SIT 2026 – kontraktové úlohy IT",
            "source": f"Filtrovaný audit PHU/SIT · stav 01–{months_loaded:02d}/{args.year}",
            "sourceFile": args.xlsx.name,
            "period": period_label(months_loaded, args.year),
            "year": args.year,
            "monthsLoaded": months_loaded,
            "method": (
                "Rozpočty a identita úloh sú zachované z existujúceho manažérskeho snapshotu. "
                "Vyčerpané sumy a mesačné čerpanie sú nanovo agregované z riadkov Filtrovaného auditu "
                "podľa stĺpcov Úloha, Suma a Mesiac/obdobie. Mesiace po poslednom načítanom mesiaci "
                "nie sú interpretované ako nulové čerpanie."
            ),
            "sourceRows": sum(row_counts.values()),
            "sourceRowsByTask": {code: row_counts[code] for code in TARGET_TASKS},
        },
        "tasks": tasks,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Written {args.output}")
    print(f"Period: {result['meta']['period']} ({months_loaded} months)")
    for task in tasks:
        print(f"Task {task['code']}: spent={task['spent']:.2f}, remaining={task['remaining']:.2f}, rows={row_counts[task['code']]}")
    print(f"Total spent: {sum(task['spent'] for task in tasks):.2f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
